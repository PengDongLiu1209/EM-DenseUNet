import os
import time
from scipy import ndimage
os.environ['KMP_DUPLICATE_LIB_OK'] = 'True'
from loss import *
from tqdm import tqdm
import torch
import numpy as np
import matplotlib.pyplot as plt
import my_confuse_matrix as my_confuse_matrix
import cv2
from PIL import Image
import torch.backends.cudnn as cudnn
import random
from typing import Dict, List, Optional
import numpy as np
import pandas as pd
from tqdm import tqdm
import torch
import os
from scipy.spatial import distance
def set_seed(seed):
    # for hash
    os.environ['PYTHONHASHSEED'] = str(seed)
    # for python and numpy
    random.seed(seed)
    np.random.seed(seed)
    # for cpu gpu
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    
    # for cudnn
    cudnn.benchmark = False
    cudnn.deterministic = True


def get_optimizer(config, model):
    assert config.opt in ['Adadelta', 'Adagrad', 'Adam', 'AdamW', 'Adamax', 'ASGD', 'RMSprop', 'Rprop', 'SGD'], 'Unsupported optimizer!'

    if config.opt == 'Adadelta':
        return torch.optim.Adadelta(
            model.parameters(),
            lr = config.lr,
            rho = config.rho,
            eps = config.eps,
            weight_decay = config.weight_decay
        )
    elif config.opt == 'Adagrad':
        return torch.optim.Adagrad(
            model.parameters(),
            lr = config.lr,
            lr_decay = config.lr_decay,
            eps = config.eps,
            weight_decay = config.weight_decay
        )
    elif config.opt == 'Adam':
        return torch.optim.Adam(
            model.parameters(),
            lr = config.lr,
            betas = config.betas,
            eps = config.eps,
            weight_decay = config.weight_decay,
            amsgrad = config.amsgrad
        )
    elif config.opt == 'AdamW':
        return torch.optim.AdamW(
            model.parameters(),
            lr = config.lr,
            betas = config.betas,
            eps = config.eps,
            weight_decay = config.weight_decay,
            amsgrad = config.amsgrad
        )
    elif config.opt == 'Adamax':
        return torch.optim.Adamax(
            model.parameters(),
            lr = config.lr,
            betas = config.betas,
            eps = config.eps,
            weight_decay = config.weight_decay
        )
    elif config.opt == 'ASGD':
        return torch.optim.ASGD(
            model.parameters(),
            lr = config.lr,
            lambd = config.lambd,
            alpha  = config.alpha,
            t0 = config.t0,
            weight_decay = config.weight_decay
        )
    elif config.opt == 'RMSprop':
        return torch.optim.RMSprop(
            model.parameters(),
            lr = config.lr,
            momentum = config.momentum,
            alpha = config.alpha,
            eps = config.eps,
            centered = config.centered,
            weight_decay = config.weight_decay
        )
    elif config.opt == 'Rprop':
        return torch.optim.Rprop(
            model.parameters(),
            lr = config.lr,
            etas = config.etas,
            step_sizes = config.step_sizes,
        )
    elif config.opt == 'SGD':
        return torch.optim.SGD(
            model.parameters(),
            lr = config.lr,
            momentum = config.momentum,
            weight_decay = config.weight_decay,
            dampening = config.dampening,
            nesterov = config.nesterov
        )
    else: # default opt is SGD
        return torch.optim.SGD(
            model.parameters(),
            lr = 0.01,
            momentum = 0.9,
            weight_decay = 0.05,
        )


def get_scheduler(config, optimizer):
    assert config.sch in ['StepLR', 'MultiStepLR', 'ExponentialLR', 'CosineAnnealingLR', 'ReduceLROnPlateau',
                        'CosineAnnealingWarmRestarts', 'WP_MultiStepLR', 'WP_CosineLR'], 'Unsupported scheduler!'
    if config.sch == 'StepLR':
        scheduler = torch.optim.lr_scheduler.StepLR(
            optimizer,
            step_size = config.step_size,
            gamma = config.gamma,
            last_epoch = config.last_epoch
        )
    elif config.sch == 'MultiStepLR':
        scheduler = torch.optim.lr_scheduler.MultiStepLR(
            optimizer,
            milestones = config.milestones,
            gamma = config.gamma,
            last_epoch = config.last_epoch
        )
    elif config.sch == 'ExponentialLR':
        scheduler = torch.optim.lr_scheduler.ExponentialLR(
            optimizer,
            gamma = config.gamma,
            last_epoch = config.last_epoch
        )
    elif config.sch == 'CosineAnnealingLR':
        scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(
            optimizer,
            T_max = config.T_max,
            eta_min = config.eta_min,
            last_epoch = config.last_epoch
        )
    elif config.sch == 'ReduceLROnPlateau':
        scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(
            optimizer, 
            mode = config.mode, 
            factor = config.factor, 
            patience = config.patience, 
            threshold = config.threshold, 
            threshold_mode = config.threshold_mode, 
            cooldown = config.cooldown, 
            min_lr = config.min_lr, 
            eps = config.eps
        )
    elif config.sch == 'CosineAnnealingWarmRestarts':
        scheduler = torch.optim.lr_scheduler.CosineAnnealingWarmRestarts(
            optimizer,
            T_0 = config.T_0,
            T_mult = config.T_mult,
            eta_min = config.eta_min,
            last_epoch = config.last_epoch
        )
    elif config.sch == 'WP_MultiStepLR':
        lr_func = lambda epoch: epoch / config.warm_up_epochs if epoch <= config.warm_up_epochs else config.gamma**len(
                [m for m in config.milestones if m <= epoch])
        scheduler = torch.optim.lr_scheduler.LambdaLR(optimizer, lr_lambda=lr_func)
    elif config.sch == 'WP_CosineLR':
        lr_func = lambda epoch: epoch / config.warm_up_epochs if epoch <= config.warm_up_epochs else 0.5 * (
                math.cos((epoch - config.warm_up_epochs) / (config.epochs - config.warm_up_epochs) * math.pi) + 1)
        scheduler = torch.optim.lr_scheduler.LambdaLR(optimizer, lr_lambda=lr_func)

    return scheduler






def save_segmentation_result(segmented_mask, out_path, num_classes, count):
    # Ensure tensor is on CPU and converted to NumPy array
    if isinstance(segmented_mask, torch.Tensor):
        segmented_mask = segmented_mask.cpu().numpy()

    # Debugging: Print original shape and unique values
    print(f"Original shape: {segmented_mask.shape}")
    print("Unique values in segmented_mask before processing:", np.unique(segmented_mask))

    # Remove batch dimension if it exists
    if segmented_mask.shape[0] == 1:  
        segmented_mask = segmented_mask[0]  # Now shape is (C, H, W)

    # Ensure it's single-channel
    if len(segmented_mask.shape) == 3:  
        segmented_mask = segmented_mask[0]  # Take the first channel (H, W)

    # Debugging: Print shape after processing
    print(f"Processed shape: {segmented_mask.shape}")

    # Define grayscale values
    if num_classes == 1:
        grayscale_values = np.array([255], dtype=np.uint8)
    else:
        grayscale_values = np.linspace(65, 255, num_classes, dtype=np.uint8)

    print("Grayscale values:", grayscale_values)

    # Map segmentation classes to grayscale values
    gray_mask = np.zeros_like(segmented_mask, dtype=np.uint8)
    for class_idx in range(num_classes):
        gray_mask[segmented_mask == class_idx] = grayscale_values[class_idx]

    # Debugging: Check unique values in the final mask
    print("Gray mask unique values before saving:", np.unique(gray_mask))

    # Convert to PIL Image
    img = Image.fromarray(gray_mask)

    # Save image
    filename = f"{count}.png"
    op = os.path.join(out_path, filename)
    img.save(op, format="PNG")

    print(f"Saved segmentation result to {op}")



# 获取多分割的前景像素点，并保存在txt文件中
def compute_gray(root):
    root = root           # 训练mask的路径
    masks_path = [os.path.join(root,i) for i in os.listdir(root)]
    gray = []           # 前景像素点
    for i in masks_path:
        img = Image.open(i)
        img = np.unique(img)        # 获取mask的灰度值
        for j in img:
            if j not in gray:
                # print(i,j)
                gray.append(j)

    with open('./data/grayList.txt','w') as f:
        gray.sort()     # 灰度值从小到大排序
        for i in gray:
            f.write(str(i))
            f.write('\n')
    return len(gray)
import torch.nn.functional as F
from torch.nn import CrossEntropyLoss

# 训练一个 epoch
def train_one_epoch(model, optim, train_loader, test_loader, device, loss_fuc, num_classes, validate=False):
    train_confmat = my_confuse_matrix.ConfusionMatrix(num_classes=num_classes)
    test_confmat = my_confuse_matrix.ConfusionMatrix(num_classes=num_classes)

    criterion = loss_fuc
    model.train()
    train_running_loss = 0.0
    train_num = 0

    # ======== 训练循环（无多余 print）========
    for train_image, train_target in tqdm(train_loader, leave=False, desc="Train"):
        train_image, train_target = train_image.to(device), train_target.to(device)

        output = model(train_image)  

        if isinstance(criterion, CrossEntropyLoss):
            # CE：直接 logits + int64 标签
            loss = criterion(output, train_target)
        else:
            # 例如 BceDiceLoss：需要 prob + one-hot
            prob = torch.softmax(output, dim=1)                      # [B,C,H,W]
            one_hot = F.one_hot(train_target, num_classes)           # [B,H,W,C]
            one_hot = one_hot.permute(0, 3, 1, 2).float()            # [B,C,H,W]
            loss = criterion(prob, one_hot)

        optim.zero_grad()
        loss.backward()
        optim.step()

        train_num += train_image.size(0)
        train_running_loss += loss.item()

        preds = torch.argmax(output, dim=1)
        
        # 记录推理时间和预测概率用于MAP和FPS计算
        # 在训练过程中，我们测量前向传播的时间作为推理时间
        start_time = time.time()
        with torch.inference_mode():
            # 重新进行一次前向传播来测量纯推理时间
            inference_output = model(train_image)
        inference_time = time.time() - start_time
        
        train_probs = torch.softmax(output, dim=1)
        
        train_confmat.update(preds, train_target, pred_probs=train_probs, inference_time=inference_time)

    train_miou_str = str(train_confmat).split('\n')[1]
    train_miou = float(train_miou_str.split(': ')[1])

    # ======== 可选验证（validate=False 时跳过）========
    test_loss = 0.0
    test_num = 0
    test_miou = 0.0
    if validate:
        model.eval()
        hd95_gather = []  # 收集每图像的每类 hd95
        with torch.inference_mode():  # 比 no_grad 更快
            for test_image, test_target in tqdm(test_loader, leave=False, desc="Val"):
                test_image, test_target = test_image.to(device), test_target.to(device)
                output = model(test_image)

                # 验证损失：和训练分支一致的 CE / BCDE 分流
                if isinstance(criterion, CrossEntropyLoss):
                    loss = criterion(output, test_target)   # <-- 修正：test_target
                else:
                    prob = torch.softmax(output, dim=1)
                    one_hot = F.one_hot(test_target, num_classes).permute(0,3,1,2).float()
                    loss = criterion(prob, one_hot)

                test_num += test_image.size(0)
                test_loss += loss.item()

                preds = torch.argmax(output, dim=1)
                
                # 记录推理时间和预测概率用于MAP和FPS计算
                # 在验证过程中，我们测量前向传播的时间作为推理时间
                start_time = time.time()
                with torch.inference_mode():
                    # 重新进行一次前向传播来测量纯推理时间
                    inference_output = model(test_image)
                inference_time = time.time() - start_time
                
                test_probs = torch.softmax(output, dim=1)
                
                test_confmat.update(preds, test_target, pred_probs=test_probs, inference_time=inference_time)

                # === 新增：逐图像/逐类 hd95（在 CPU 上做）===
                preds_np   = preds.detach().cpu().numpy()
                targets_np = test_target.detach().cpu().numpy()
                bs = preds_np.shape[0]
                for i in range(bs):
                    hd = compute_hd95_per_class(
                        preds_np[i], targets_np[i],
                        num_classes=num_classes, ignore_index=255, spacing=None
                    )
                    hd95_gather.append(hd)

        # 原有 mIoU 提取
        test_miou_str = str(test_confmat).split('\n')[1]
        test_miou = float(test_miou_str.split(': ')[1])

        # 聚合 hd95（每类平均）
        if len(hd95_gather) > 0:
            hd95_per_class = np.nanmean(np.vstack(hd95_gather), axis=0)
        else:
            hd95_per_class = np.full(num_classes, np.nan, dtype=float)

        hd95_strs = ['{:.4f}'.format(x) if not np.isnan(x) else 'nan' for x in hd95_per_class.tolist()]
        val_confmat_str = str(test_confmat) + f"hd95: {hd95_strs}\n"
    else:
        val_confmat_str = None

    lr = optim.param_groups[0]["lr"]
    return (
        train_running_loss / train_num,
        (test_loss / test_num) if validate else None,
        lr,
        train_miou,
        (test_miou if validate else None),
        val_confmat_str,  # 注意：现在返回的是拼了 hd95 的字符串
    )
# 计算性能指标
def evaluate(model, train_loader, test_loader, device, num):
    model.eval()
    train_confmat = my_confuse_matrix.ConfusionMatrix(num_classes=num)
    test_confmat  = my_confuse_matrix.ConfusionMatrix(num_classes=num)

    # 用 inference_mode 替换 no_grad（更快）
    with torch.inference_mode():
        for train_image, train_target in tqdm(train_loader, leave=False):  # 可选 leave=False
            train_image, train_target = train_image.to(device), train_target.to(device)
            
            # 记录推理时间
            start_time = time.time()
            train_output = model(train_image)
            inference_time = time.time() - start_time
            
            # 获取预测概率和预测类别
            train_probs = torch.softmax(train_output, dim=1)
            train_output = torch.argmax(train_output, dim=1)
            
            # 更新混淆矩阵，包含MAP和FPS所需的数据
            train_confmat.update(train_output, train_target, pred_probs=train_probs, inference_time=inference_time)

        train_miou_str = str(train_confmat).split('\n')[1]
        train_miou = float(train_miou_str.split(': ')[1])

        for test_image, test_target in tqdm(test_loader, leave=False):
            test_image, test_target = test_image.to(device), test_target.to(device)
            
            # 记录推理时间
            start_time = time.time()
            test_output = model(test_image)
            inference_time = time.time() - start_time
            
            # 获取预测概率和预测类别
            test_probs = torch.softmax(test_output, dim=1)
            test_output = torch.argmax(test_output, dim=1)
            
            # 更新混淆矩阵，包含MAP和FPS所需的数据
            test_confmat.update(test_output, test_target, pred_probs=test_probs, inference_time=inference_time)

        test_miou_str = str(test_confmat).split('\n')[1]
        test_miou = float(test_miou_str.split(': ')[1])

    return train_miou, test_miou, str(test_confmat)

import numpy as np
import torch

def evaluate_with_hd95(model, loader, device, num_classes, ignore_index=255, spacing=None):
    from my_confuse_matrix import ConfusionMatrix
    confmat = ConfusionMatrix(num_classes=num_classes, ignore_index=ignore_index)

    hd95_all = []  # [num_images, num_classes]

    model.eval()
    with torch.inference_mode():
        for images, targets in loader:
            images  = images.to(device)
            targets = targets.to(device)
            logits  = model(images)
            preds   = torch.argmax(logits, dim=1)

            confmat.update(preds, targets)  # 混淆矩阵（快）

            # 逐图像计算 hd95（放 CPU）
            preds_np  = preds.detach().cpu().numpy()
            targets_np= targets.detach().cpu().numpy()
            for i in range(preds_np.shape[0]):
                hd = compute_hd95_per_class(preds_np[i], targets_np[i], num_classes,
                                            ignore_index=ignore_index, spacing=spacing)
                hd95_all.append(hd)

    # 指标
    acc_global, recall, precision, iou, dice = confmat.compute()
    miou  = iou.mean().item()
    mdice = dice.mean().item()
    hd95_per_class = np.nanmean(np.vstack(hd95_all), axis=0) if len(hd95_all) > 0 else np.full(num_classes, np.nan)
    hd95_macro     = float(np.nanmean(hd95_per_class))

    # 如果你下游仍需键名 "hd95"，这里可以保持兼容：
    log = (
        'macc: {:.4f}\n'
        'mIoU: {:.4f}\n'
        'precision: {}\n'
        'recall: {}\n'
        'IoU: {}\n'
        'Dice: {}\n'
        'mDice: {:.4f}\n'
        'hd95: {}\n'   # 实际上是 hd95
    ).format(
        acc_global.item(),
        miou,
        ['{:.4f}'.format(x) for x in precision.tolist()],
        ['{:.4f}'.format(x) for x in recall.tolist()],
        ['{:.4f}'.format(x) for x in iou.tolist()],
        ['{:.4f}'.format(x) for x in dice.tolist()],
        mdice,
        ['{:.4f}'.format(x) if not np.isnan(x) else 'nan' for x in hd95_per_class.tolist()]
    )

    return miou, log, hd95_per_class, hd95_macro


def test(model, test_loader, device, num):
    model.eval()
    test_confmat = my_confuse_matrix.ConfusionMatrix(num_classes=num)

    with torch.inference_mode():
        for test_image, test_target in tqdm(test_loader, leave=False):
            test_image, test_target = test_image.to(device), test_target.to(device)
            
            # 记录推理时间
            start_time = time.time()
            test_output = model(test_image)
            inference_time = time.time() - start_time
            
            # 获取预测概率和预测类别
            test_probs = torch.softmax(test_output, dim=1)
            test_output = torch.argmax(test_output, dim=1)
            
            # 更新混淆矩阵，包含MAP和FPS所需的数据
            test_confmat.update(test_output, test_target, pred_probs=test_probs, inference_time=inference_time)

        test_miou_str = str(test_confmat).split('\n')[1]
        test_miou = float(test_miou_str.split(': ')[1])

    return test_miou, str(test_confmat)


def test_out(model, test_loader, device, num, out_path):
    model.eval()
    test_confmat = my_confuse_matrix.ConfusionMatrix(num_classes=num)
    count = 1

    with torch.inference_mode():
        for test_image, test_target in tqdm(test_loader, leave=False):
            test_image, test_target = test_image.to(device), test_target.to(device)

            logits = model(test_image)
            pred   = torch.argmax(logits, dim=1)  # [B,H,W] 的类别图

            # 保存时传“类别图”而不是 logits（更快、更省显存、更正确）
            save_segmentation_result(segmented_mask=pred, out_path=out_path,
                                     num_classes=num, count=count)
            count += 1

            test_confmat.update(pred, test_target)

        test_miou_str = str(test_confmat).split('\n')[1]
        test_miou = float(test_miou_str.split(': ')[1])

    return test_miou, str(test_confmat)



# 可视化数据,只展示2个
def plot(data_loader,mean=(0.5, 0.5, 0.5), std=(0.5, 0.5, 0.5)):
    plt.figure(figsize=(12, 8))
    imgs, labels = data_loader
    print('images:',imgs.shape,imgs.dtype)                  # torch.Size([batch, 3, 96, 96]) torch.float32
    print('labels:',labels.shape,labels.dtype)              # torch.Size([batch, 96, 96]) torch.int64
    print('classes:',np.unique(labels))                      # 0 1 255 只包含 0 1 2...255(255为预处理填充的部分)
    save_dir = './see'
    for i, (x, y) in enumerate(zip(imgs[:2], labels[:2])):
        x = np.transpose(x.numpy(), (1, 2, 0))
        x[:, :, 0] = x[:, :, 0] * std[0] + mean[0]  # 去 normalization
        x[:, :, 1] = x[:, :, 1] * std[1] + mean[1]
        x[:, :, 2] = x[:, :, 2] * std[2] + mean[2]
        y = y.numpy()

        plt.subplot(2, 2, i + 1)
        plt.imshow(x)

        plt.subplot(2, 2, i + 3)
        plt.imshow(y)
            # 保存图像到指定文件夹
        plt.savefig(os.path.join(save_dir, f'image_{i}.png'))
        plt.close() 


# 绘制学习率衰减过程
def plot_lr_decay(scheduler, optimizer, epochs, LR_path):
    # 备份真实状态
    sd_sched = scheduler.state_dict()
    sd_opt   = optimizer.state_dict()

    y = []
    for _ in range(epochs):
        scheduler.step()
        y.append(optimizer.param_groups[0]['lr'])

    plt.plot(y, '.-', label='LambdaLR')
    plt.xlabel('epoch'); plt.ylabel('LR')
    plt.tight_layout()
    plt.savefig(LR_path, dpi=300)

    # 恢复真实状态（关键）
    scheduler.load_state_dict(sd_sched)
    optimizer.load_state_dict(sd_opt)



# 绘制loss 和 iou曲线
def plt_loss_iou(train_loss,test_loss,train_iou,test_iou,SAVE_IOU_PATH):
    plt.figure(figsize=(12,8))
    plt.subplot(1,2,1)
    plt.plot(train_loss,label='train loss',linestyle='-',color='g')
    plt.plot(test_loss,label='test loss',linestyle='-.',color='r')
    plt.title('loss curve')
    plt.legend()

    plt.subplot(1,2,2)
    plt.plot(train_iou,label='train mean iou',linestyle='-',color='g')
    plt.plot(test_iou,label='test mean iou',linestyle='-.',color='r')
    plt.title('mean iou curve')
    plt.legend()

    plt.savefig(SAVE_IOU_PATH,dpi=300)
# ==== 新增：逐图像 IoU/Dice 计算（忽略 index=255）====
import pandas as pd  # 若你环境没有 pandas，可用内置 csv 替代，见后文可选写法
from typing import Dict, List, Optional

def _single_image_iou_dice(pred_np, tgt_np, num_classes: int, ignore_index: int = 255):
    """
    pred_np, tgt_np: [H, W] 的 numpy 数组（整型类别索引）
    返回： iou_per_class, dice_per_class（长度为 num_classes，可能含 NaN）
    """
    mask = (tgt_np != ignore_index)
    pred = pred_np[mask]
    tgt = tgt_np[mask]

    iou_per_class = np.full(num_classes, np.nan, dtype=float)
    dice_per_class = np.full(num_classes, np.nan, dtype=float)

    for c in range(num_classes):
        pred_c = (pred == c)
        tgt_c  = (tgt  == c)
        inter = np.logical_and(pred_c, tgt_c).sum()
        union = np.logical_or (pred_c, tgt_c).sum()
        denom_dice = pred_c.sum() + tgt_c.sum()

        iou_per_class[c]  = (inter / union) if union > 0 else np.nan
        dice_per_class[c] = (2 * inter / denom_dice) if denom_dice > 0 else np.nan

    return iou_per_class, dice_per_class


# ==== 新增：测试并导出逐图像指标 ====



def _single_image_iou_dice(pred_np, tgt_np, num_classes: int, ignore_index: int = 255):
    mask = (tgt_np != ignore_index)
    pred = pred_np[mask]
    tgt = tgt_np[mask]

    iou_per_class = np.full(num_classes, np.nan, dtype=float)
    dice_per_class = np.full(num_classes, np.nan, dtype=float)

    for c in range(num_classes):
        pred_c = (pred == c)
        tgt_c = (tgt == c)
        inter = np.logical_and(pred_c, tgt_c).sum()
        union = np.logical_or(pred_c, tgt_c).sum()
        denom_dice = pred_c.sum() + tgt_c.sum()

        iou_per_class[c] = (inter / union) if union > 0 else np.nan
        dice_per_class[c] = (2 * inter / denom_dice) if denom_dice > 0 else np.nan

    return iou_per_class, dice_per_class



import GeodisTK

def get_edge_points(img):
    """
    get edge points of a binary segmentation result
    """
    dim = len(img.shape)
    if (dim == 2):
        strt = ndimage.generate_binary_structure(2, 1)
    else:
        strt = ndimage.generate_binary_structure(3, 1)  # 三维结构元素，与中心点相距1个像素点的都是邻域
    ero = ndimage.morphology.binary_erosion(img, strt)
    edge = np.asarray(img, np.uint8) - np.asarray(ero, np.uint8)
    return edge

def binary_hausdorff95(s, g, spacing=None):
    """
    get the hausdorff distance between a binary segmentation and the ground truth
    inputs:
        s: a 3D or 2D binary image for segmentation
        g: a 2D or 2D binary image for ground truth
        spacing: a list for image spacing, length should be 3 or 2
    """
    s_edge = get_edge_points(s)
    g_edge = get_edge_points(g)
    image_dim = len(s.shape)
    assert (image_dim == len(g.shape))
    if (spacing == None):
        spacing = [1.0] * image_dim
    else:
        assert (image_dim == len(spacing))
    img = np.zeros_like(s)
    if (image_dim == 2):
        s_dis = GeodisTK.geodesic2d_raster_scan(img, s_edge, 0.0, 2)
        g_dis = GeodisTK.geodesic2d_raster_scan(img, g_edge, 0.0, 2)
    elif (image_dim == 3):
        s_dis = GeodisTK.geodesic3d_raster_scan(img, s_edge, spacing, 0.0, 2)
        g_dis = GeodisTK.geodesic3d_raster_scan(img, g_edge, spacing, 0.0, 2)

    dist_list1 = s_dis[g_edge > 0]
    dist_list1 = sorted(dist_list1)
    dist1 = dist_list1[int(len(dist_list1) * 0.95)]
    dist_list2 = g_dis[s_edge > 0]
    dist_list2 = sorted(dist_list2)
    dist2 = dist_list2[int(len(dist_list2) * 0.95)]
    return max(dist1, dist2)


# 平均表面距离
def binary_assd(s, g, spacing=None):
    """
    get the average symetric surface distance between a binary segmentation and the ground truth
    inputs:
        s: a 3D or 2D binary image for segmentation
        g: a 2D or 2D binary image for ground truth
        spacing: a list for image spacing, length should be 3 or 2
    """
    s_edge = get_edge_points(s)
    g_edge = get_edge_points(g)
    image_dim = len(s.shape)
    assert (image_dim == len(g.shape))
    if (spacing == None):
        spacing = [1.0] * image_dim
    else:
        assert (image_dim == len(spacing))
    img = np.zeros_like(s)
    if (image_dim == 2):
        s_dis = GeodisTK.geodesic2d_raster_scan(img, s_edge, 0.0, 2)
        g_dis = GeodisTK.geodesic2d_raster_scan(img, g_edge, 0.0, 2)
    elif (image_dim == 3):
        s_dis = GeodisTK.geodesic3d_raster_scan(img, s_edge, spacing, 0.0, 2)
        g_dis = GeodisTK.geodesic3d_raster_scan(img, g_edge, spacing, 0.0, 2)

    ns = s_edge.sum()
    ng = g_edge.sum()
    s_dis_g_edge = s_dis * g_edge
    g_dis_s_edge = g_dis * s_edge
    assd = (s_dis_g_edge.sum() + g_dis_s_edge.sum()) / (ns + ng)
    return assd



# 放在 utils.py 或单独 utils_metrics.py 里
import numpy as np
from medpy.metric.binary import hd95

def compute_hd95_per_class(pred_np: np.ndarray,
                           tgt_np:  np.ndarray,
                           num_classes: int,
                           ignore_index: int = 255,
                           spacing=None) -> np.ndarray:
    """
    pred_np, tgt_np: [H,W] int，类别索引
    返回: 长度=num_classes 的向量，每类 hd95（单位=像素或 spacing 单位）。无有效像素则返回 np.nan
    """
    # 忽略无效像素
    valid = (tgt_np != ignore_index)
    pred = pred_np.copy()
    tgt  = tgt_np.copy()
    pred[~valid] = 0
    tgt[~valid]  = 0

    out = np.full(num_classes, np.nan, dtype=float)
    for c in range(num_classes):
        pred_c = (pred == c).astype(np.uint8)
        tgt_c  = (tgt  == c).astype(np.uint8)
        if pred_c.any() and tgt_c.any():
            try:
                out[c] = hd95(pred_c, tgt_c, voxelspacing=spacing)
            except Exception:
                out[c] = np.nan
    return out






from medpy.metric.binary import hd95
def _single_image_hd95(pred_np, tgt_np, num_classes: int, ignore_index: int = 255, spacing=None):
    """
    pred_np, tgt_np: [H, W] 的整型类别图
    spacing: None 或 (sy, sx) / (sz, sy, sx)，传给 hd95 的 voxelspacing
    返回：每类 hd95（长度 num_classes，可能含 NaN）
    """
    hd95_per_class = np.full(num_classes, np.nan, dtype=float)

    # 有效像素（非 ignore 区域）
    valid = (tgt_np != ignore_index)

    for c in range(num_classes):
        # 仅在有效区域上取该类二值掩码
        pred_c = (pred_np == c) & valid
        tgt_c  = (tgt_np == c)  & valid

        # 当任一为空，hd95 没定义，保留 NaN（和你现在的逻辑一致）
        if np.count_nonzero(pred_c) == 0 or np.count_nonzero(tgt_c) == 0:
            continue

        try:
            # 如果用的是 medpy.metric.binary.hd95，可以这样传 spacing
            if spacing is None:
                hd95_per_class[c] = binary_hausdorff95(pred_c, tgt_c)
            else:
                hd95_per_class[c] = binary_hausdorff95(pred_c, tgt_c, spacing=spacing)
        except Exception:
            hd95_per_class[c] = np.nan

    return hd95_per_class


def colorize_mask(mask: np.ndarray, num_classes: int) -> np.ndarray:
    """
    将灰度掩码 mask 按类别着色，返回 RGB 彩图。
    mask: (H, W)，每个像素为类别标签 [0, num_classes-1]
    return: (H, W, 3) 彩色掩码图像
    """
    cmap = plt.get_cmap('tab20', num_classes)
    h, w = mask.shape
    color_mask = np.zeros((h, w, 3), dtype=np.uint8)

    for cls in range(num_classes):
        r, g, b, _ = cmap(cls)
        color_mask[mask == cls] = (int(r * 255), int(g * 255), int(b * 255))

    return color_mask


def overlay_mask_on_image_multiclass(image: np.ndarray, mask: np.ndarray, num_classes: int, alpha=0.5, color_map=None):
    if color_map is None:
        np.random.seed(42)
        color_map = [tuple(np.random.randint(0, 255, size=3).tolist()) for _ in range(num_classes)]
    
    color_mask = np.zeros_like(image)
    for c in range(num_classes):
        color = color_map[c]
        color_mask[mask == c] = color

    overlay = cv2.addWeighted(image, 1 - alpha, color_mask, alpha, 0)
    return overlay

def add_legend(image: np.ndarray, class_names: list, color_map: list, position=(10, 10)):
    legend_img = image.copy()
    x, y = position
    font = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 0.5
    spacing = 20

    for i, (name, color) in enumerate(zip(class_names, color_map)):
        cv2.rectangle(legend_img, (x, y + i * spacing), (x + 15, y + i * spacing + 15), color[::-1], -1)
        cv2.putText(legend_img, name, (x + 20, y + i * spacing + 12), font, font_scale, (255, 255, 255), thickness=1)

    return legend_img


from scipy.stats import mannwhitneyu

def test_with_per_image(
    model,
    test_loader,
    device,
    num_classes: int,
    save_csv_path: Optional[str] = None,
    ignore_index: int = 255,
    save_vis_dir: Optional[str] = None,
):
    model.eval()
    class_names = ['Background', 'Class1', 'Class2', 'Class3']
    gt_colors = [(0, 0, 255), (128, 0, 128), (255, 0, 255), (255, 128, 128)]  # Red hues: GT
    pred_colors = [(0, 255, 0), (0, 255, 255), (0, 128, 255), (128, 255, 128)]  # Green hues: Pred

    test_confmat = my_confuse_matrix.ConfusionMatrix(num_classes=num_classes)

    image_ids: List[str] = []
    iou_macro_list: List[float] = []
    dice_macro_list: List[float] = []
    hd95_macro_list: List[float] = []

    iou_per_class_all: List[np.ndarray] = []
    dice_per_class_all: List[np.ndarray] = []
    hd95_per_class_all: List[np.ndarray] = []

    g_idx = 0
    with torch.inference_mode():
        for batch in tqdm(test_loader):
            if len(batch) >= 2:
                test_image, test_target = batch[0].to(device), batch[1].to(device)
                extra = batch[2] if len(batch) >= 3 else None
            else:
                raise ValueError("test_loader should return at least (image, target)")

            # Measure inference time
            start_time = time.time()
            logits = model(test_image)
            end_time = time.time()
            inference_time = end_time - start_time
            
            pred = torch.argmax(logits, dim=1)
            
            # Calculate prediction probabilities using softmax
            pred_probs = torch.softmax(logits, dim=1)

            test_confmat.update(pred, test_target, pred_probs=pred_probs, inference_time=inference_time)

            pred_np = pred.cpu().numpy()
            tgt_np = test_target.cpu().numpy()

            bs = pred_np.shape[0]
            for i in range(bs):
                # Image ID assignment
                if extra is not None:
                    if isinstance(extra, (list, tuple)) and len(extra) == bs:
                        iid = os.path.basename(str(extra[i]))
                    else:
                        iid = str(extra)
                else:
                    iid = f"{g_idx:06d}"
                g_idx += 1
                image_ids.append(iid)

                # IoU and Dice calculation
                iou_c, dice_c = _single_image_iou_dice(pred_np[i], tgt_np[i], num_classes, ignore_index)
                hd95_c = compute_hd95_per_class(pred_np[i], tgt_np[i], num_classes, ignore_index=ignore_index)

                iou_macro  = np.nanmean(iou_c)
                dice_macro = np.nanmean(dice_c)
                hd95_macro = np.nanmean(hd95_c)

                iou_macro_list.append(iou_macro)
                dice_macro_list.append(dice_macro)
                hd95_macro_list.append(hd95_macro)

                iou_per_class_all.append(iou_c)
                dice_per_class_all.append(dice_c)
                hd95_per_class_all.append(hd95_c)

                # Mann-Whitney U Test for comparing GT vs Pred
                # Perform Mann-Whitney U Test on IoU, Dice, and HD95 scores per class
                p_iou = mannwhitneyu(iou_c, iou_c, alternative='two-sided')[1]
                p_dice = mannwhitneyu(dice_c, dice_c, alternative='two-sided')[1]
                p_hd95 = mannwhitneyu(hd95_c, hd95_c, alternative='two-sided')[1]

                if p_iou < 0.05:
                    print(f"Statistically significant difference in IoU for image {iid} (p = {p_iou:.4f})")
                if p_dice < 0.05:
                    print(f"Statistically significant difference in Dice for image {iid} (p = {p_dice:.4f})")
                if p_hd95 < 0.05:
                    print(f"Statistically significant difference in HD95 for image {iid} (p = {p_hd95:.4f})")

                # Saving visualizations
                if save_vis_dir is not None:
                    os.makedirs(save_vis_dir, exist_ok=True)
                    save_name_base = os.path.splitext(iid)[0]

                    # Save original image
                    img_np = test_image[i].detach().cpu().numpy()
                    img_np = (img_np * 255).astype(np.uint8)
                    img_np = np.transpose(img_np, (1, 2, 0))

                    # Save predicted & ground truth masks
                    pred_mask = pred_np[i].astype(np.uint8)
                    gt_mask = tgt_np[i].astype(np.uint8)

                    # 生成预测掩码时忽略背景类
                    pred_mask_no_bg = np.where(pred_mask == 0, 255, pred_mask)  # 将背景类设为255或忽略

                    # 生成地面真值（GT）掩码时，确保忽略背景类
                    gt_mask_no_bg = np.where(gt_mask == 0, 255, gt_mask)  # 将背景类设为255或忽略

                    # 保存原图
                    img_np = test_image[i].detach().cpu().numpy()
                    img_np = (img_np * 255).astype(np.uint8)
                    img_np = np.transpose(img_np, (1, 2, 0))  # 转换为HWC格式

                    # 保存预测掩码
                    pred_color = colorize_mask(pred_mask_no_bg, num_classes)
                    Image.fromarray(pred_color).save(os.path.join(save_vis_dir, f"{save_name_base}_pred.png"))

                    gt_mask_no_bg_transparent = np.where(gt_mask == 0, 255, gt_mask)  # 背景类设为255
                    overlay_gt = overlay_mask_on_image_multiclass(
                        img_np,
                        gt_mask_no_bg_transparent,
                        num_classes,
                        alpha=0.5,
                        color_map=pred_colors
                    )
                    Image.fromarray(overlay_gt).save(os.path.join(save_vis_dir, f"{save_name_base}_overlay_gt.png"))
                    # 原图+预测掩码（绿色）
                    overlay_pred = overlay_mask_on_image_multiclass(img_np, pred_mask_no_bg, num_classes, color_map=pred_colors, alpha=0.5)
                    Image.fromarray(overlay_pred).save(os.path.join(save_vis_dir, f"{save_name_base}_overlay_pred.png"))

                    # 原图+预测（绿色）和GT掩码（红色）
                    overlay_both = overlay_mask_on_image_multiclass(img_np, pred_mask_no_bg, num_classes, color_map=pred_colors, alpha=0.5)
                    overlay_both = overlay_mask_on_image_multiclass(overlay_both, gt_mask_no_bg, num_classes, color_map=gt_colors, alpha=0.5)

                    # 添加图例
                    legend_labels = [f"Pred-Class{i}" for i in range(1, num_classes)] + [f"GT-Class{i}" for i in range(1, num_classes)]
                    legend_colors = pred_colors[1:] + gt_colors[1:]
                    overlay_both = add_legend(overlay_both, legend_labels, legend_colors, position=(10, 10))

                    # 保存带图例的合成图
                    Image.fromarray(overlay_both).save(os.path.join(save_vis_dir, f"{save_name_base}_overlay_both.png"))


        test_miou_str = str(test_confmat).split('\n')[1]
        test_miou = float(test_miou_str.split(': ')[1])

    # Return per-image data
    per_image: Dict[str, np.ndarray] = {
        "image_id": image_ids,
        "iou_macro": np.array(iou_macro_list, dtype=float),
        "dice_macro": np.array(dice_macro_list, dtype=float),
        "hd95_macro": np.array(hd95_macro_list, dtype=float),
        "iou_per_class": np.vstack(iou_per_class_all),
        "dice_per_class": np.vstack(dice_per_class_all),
        "hd95_per_class": np.vstack(hd95_per_class_all),
    }
    if len(hd95_per_class_all) > 0:
        hd95_per_class_mean = np.nanmean(np.vstack(hd95_per_class_all), axis=0)
    else:
        hd95_per_class_mean = np.full(num_classes, np.nan, dtype=float)

    hd95_macro_overall = float(np.nanmean(hd95_per_class_mean))
    hd95_strs = ['{:.4f}'.format(x) if not np.isnan(x) else 'nan' for x in hd95_per_class_mean.tolist()]

    # 把 hd95 写进 confusion matrix 的“额外指标”
    test_confmat.set_extra(
        hd95_per_class=hd95_strs,            # ['3.9279', '8.0946', '5.6633']
        hd95_macro=f'{hd95_macro_overall:.4f}',
        hd95=hd95_strs                       # 兼容旧逻辑
    )

    # If you want to save it to CSV
    if save_csv_path is not None:
        rows = {
            "image_id": image_ids,
            "iou_macro": iou_macro_list,
            "dice_macro": dice_macro_list,
            "hd95_macro": hd95_macro_list,
        }
        for c in range(num_classes):
            rows[f"iou_c{c}"] = [v[c] for v in iou_per_class_all]
            rows[f"dice_c{c}"] = [v[c] for v in dice_per_class_all]
            rows[f"hd95_c{c}"] = [v[c] for v in hd95_per_class_all]
        df = pd.DataFrame(rows)
        os.makedirs(os.path.dirname(save_csv_path) or ".", exist_ok=True)
        df.to_csv(save_csv_path, index=False)

    return test_miou, test_confmat, per_image



def append_log_to_excel(log, model_name, categories, out_path):

    def _to_float(x):
        try: 
            return float(x) if x is not None else None
        except: 
            return None

    n = len(categories)

    # Modified _vec function to work with dictionary
    def _vec(key):
        # Handle case sensitivity
        key_lower = key.lower()
        for k in log.keys():
            if k.lower() == key_lower:
                value = log[k]
                if isinstance(value, list) and len(value) == n:
                    return value
                elif isinstance(value, str):
                    return [_to_float(value)] * n
        return [None] * n

    precision = _vec("precision")
    recall    = _vec("recall")
    iou       = _vec("iou")   # Note: case-insensitive matching
    dice      = _vec("dice")
    ap        = _vec("ap")    # 添加AP支持

    # Handle HD95: try 'hd95' first, then 'hd95_per_class'
    hd95_vec = _vec("hd95")
    if all(x is None for x in hd95_vec):
        hd95_vec = _vec("hd95_per_class")

    # Global metrics
    macc        = _to_float(log.get('macc'))
    mIoU        = _to_float(log.get('mIoU'))
    mDice       = _to_float(log.get('mDice'))
    hd95_macro  = _to_float(log.get('hd95_macro'))
    mAP         = _to_float(log.get('mAP'))    # 添加mAP支持
    fps         = _to_float(log.get('FPS'))    # 添加FPS支持

    rows = []
    for i in range(n):
        rows.append({
            "model": model_name,
            "class": categories[i],
            "macc": macc,
            "mIoU": mIoU,
            "precision": _to_float(precision[i] if i < len(precision) else None),
            "recall":    _to_float(recall[i] if i < len(recall) else None),
            "IoU":       _to_float(iou[i] if i < len(iou) else None),
            "mDice":     mDice,
            "Dice":      _to_float(dice[i] if i < len(dice) else None),
            "hd95":      _to_float(hd95_vec[i] if i < len(hd95_vec) else None),
            "hd95_macro": hd95_macro,
            "AP":        _to_float(ap[i] if i < len(ap) else None),  # 添加AP列
            "mAP":       mAP,      # 添加mAP列
            "FPS":       fps,      # 添加FPS列
        })

    df_new = pd.DataFrame(rows)

    # Fixed column order - 添加新的列
    cols = ["model","class","macc","mIoU","precision","recall","IoU","mDice","Dice","hd95","hd95_macro","AP","mAP","FPS"]
    df_new = df_new[cols]

    # Save to file (same as original)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    if out_path.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        if os.path.exists(out_path):
            book = load_workbook(out_path)
            with pd.ExcelWriter(out_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                writer.book = book
                sheet = "metrics"
                if sheet not in writer.book.sheetnames:
                    df_new.to_excel(writer, sheet_name=sheet, index=False)
                else:
                    old = pd.read_excel(out_path, sheet_name=sheet)
                    all_df = pd.concat([old, df_new], ignore_index=True)
                    writer.book.remove(writer.book[sheet])
                    writer.book.create_sheet(sheet)
                    all_df.to_excel(writer, sheet_name=sheet, index=False)
        else:
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df_new.to_excel(writer, sheet_name="metrics", index=False)
    else:
        write_header = not os.path.exists(out_path)
        df_new.to_csv(out_path, mode="a", index=False, header=write_header)